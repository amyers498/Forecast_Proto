import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from prophet import Prophet
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
import logging
import numpy as np
import re
import plotly.graph_objects as go

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Define industries and their predefined fields
INDUSTRY_FIELDS = {
    "Tech Startups": ["Cloud Costs", "API Costs", "Software Subscriptions", "R&D Expenses", "Revenue"],
    "Retail": ["Inventory Costs", "Sales Revenue", "Marketing Spend", "Store Operating Costs"],
    "Healthcare": ["Medical Supplies", "Staff Salaries", "Patient Revenue", "Insurance Claims"],
    "Hospitality": ["Food/Beverage Costs", "Staff Wages", "Utility Costs", "Room Revenue"],
    "Manufacturing": ["Raw Material Costs", "Production Costs", "Labor Costs", "Sales Revenue"],
    "E-commerce": ["Hosting Costs", "Advertising Costs", "Shipping Costs", "Online Sales Revenue"],
    "Education": ["Teacher Salaries", "Facility Costs", "Supplies Costs", "Tuition Revenue"],
    "Construction": ["Material Costs", "Labor Costs", "Equipment Rental", "Project Revenue"],
    "Non-Profit Organizations": ["Donations Revenue", "Grants", "Program Costs", "Staff Salaries"],
    "Real Estate": ["Property Acquisition Costs", "Rental Income", "Maintenance Costs", "Property Taxes"]
}

# Grok 3 Beta API setup
GROK_API_KEY = "xai-S9KY18eQlENXtmS2EaQbodZFbHBXARKD03x2ZEb2b3Zx8afusaSjxbtOLm9Twyt1WvTQJO150IorIan7"
GROK_API_URL = "https://api.x.ai/v1/chat/completions"
GROK_API_URL_FALLBACK = "https://api.x.ai/v1/completions"

# Cache for Grok 3 Beta responses (for repeatability)
if 'grok_cache' not in st.session_state:
    st.session_state.grok_cache = {}

# Define current_date at the top level for shared use
current_date = datetime.now().strftime("%Y%m%d")

def call_grok_api(prompt, cache_key=None, high_effort=False):
    if cache_key and cache_key in st.session_state.grok_cache:
        return st.session_state.grok_cache[cache_key]

    logger.info(f"Attempting to connect to primary endpoint: {GROK_API_URL}")

    session = requests.Session()
    retries = Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504])
    session.mount('https://', HTTPAdapter(max_retries=retries))

    headers = {
    "Authorization": f"Bearer {st.secrets['api']['grok_api_key']}",
    "Content-Type": "application/json"
    }
    
    payload_chat = {
        "model": "grok-3-beta",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 100
    }
    
    payload_completions = {
        "model": "grok-3-beta",
        "prompt": prompt,
        "temperature": 0.1,
        "max_tokens": 100
    }

    if high_effort:
        payload_chat["reasoning"] = {"effort": "high"}
        payload_completions["reasoning"] = {"effort": "high"}

    try:
        logger.info(f"Sending request to {GROK_API_URL} with payload: {json.dumps(payload_chat, indent=2)}")
        response = session.post(GROK_API_URL, headers=headers, json=payload_chat, timeout=30)
        response.raise_for_status()
        raw_response = response.json()
        logger.info(f"Raw API response: {json.dumps(raw_response, indent=2)}")
        if 'choices' in raw_response and len(raw_response['choices']) > 0:
            if 'message' in raw_response['choices'][0] and 'content' in raw_response['choices'][0]['message']:
                result = raw_response['choices'][0]['message']['content']
            else:
                result = "Incomplete response from API. Please try again."
        else:
            result = "No valid response from API. Please try again."
        logger.info(f"Parsed response: {result}")

        # Post-process to ensure TLDR is concise and complete sentences
        sentences = re.split(r'(?<=[.!?])\s+', result.strip())
        tldr_sentences = sentences[:2]  # Take first 1-2 sentences
        tldr = ' '.join(tldr_sentences)
        if len(sentences) > 2:
            tldr += "..."
        result = tldr if tldr else result

    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            logger.info(f"Primary endpoint failed with 404. Trying fallback endpoint: {GROK_API_URL_FALLBACK}")
            try:
                logger.info(f"Sending request to {GROK_API_URL_FALLBACK} with payload: {json.dumps(payload_completions, indent=2)}")
                response = session.post(GROK_API_URL_FALLBACK, headers=headers, json=payload_completions, timeout=30)
                response.raise_for_status()
                raw_response = response.json()
                logger.info(f"Raw API response (fallback): {json.dumps(raw_response, indent=2)}")
                if 'choices' in raw_response and len(raw_response['choices']) > 0:
                    if 'text' in raw_response['choices'][0]:
                        result = raw_response['choices'][0]['text']
                    else:
                        result = "Incomplete response from API (fallback). Please try again."
                else:
                    result = "No valid response from API (fallback). Please try again."
                logger.info(f"Parsed response (fallback): {result}")

                # Post-process for TLDR
                sentences = re.split(r'(?<=[.!?])\s+', result.strip())
                tldr_sentences = sentences[:2]
                tldr = ' '.join(tldr_sentences)
                if len(sentences) > 2:
                    tldr += "..."
                result = tldr if tldr else result

            except requests.exceptions.HTTPError as e:
                st.error(f"HTTP error from Grok 3 Beta API: {str(e)}. Both endpoints failed. Ensure your API key has access to Grok 3 Beta and the model is available.")
                st.stop()
            except Exception as e:
                st.error(f"Unexpected error when calling Grok 3 Beta API on fallback endpoint: {str(e)}.")
                st.stop()
        else:
            st.error(f"HTTP error from Grok 3 Beta API: {str(e)}. Ensure your API key is valid and the endpoint is correct.")
            st.stop()
    except requests.exceptions.ConnectionError as e:
        st.error(f"Failed to connect to Grok 3 Beta API: {str(e)}. Please check your network connection, DNS settings, or try again later.")
        st.stop()
    except requests.exceptions.Timeout:
        st.error("Request to Grok 3 Beta API timed out. The API might be down or your network is slow. Try again later.")
        st.stop()
    except Exception as e:
        st.error(f"Unexpected error when calling Grok 3 Beta API: {str(e)}.")
        st.stop()

    if cache_key:
        st.session_state.grok_cache[cache_key] = result
    return result

# Function to extrapolate regressor trends (used for metrics forecasting)
def extrapolate_regressor(historical_data, dates, future_dates, percentage_increase):
    x = np.arange(len(historical_data))
    coeffs = np.polyfit(x, historical_data, 1)  # Linear fit (degree 1)
    trend = np.poly1d(coeffs)
    
    all_dates_indices = np.arange(len(dates))
    extrapolated_values = trend(all_dates_indices)
    
    future_start_idx = len(historical_data)
    for i in range(future_start_idx, len(dates)):
        extrapolated_values[i] *= (1 + percentage_increase / 100)
    
    return extrapolated_values

# Streamlit app
st.title("Budgeting and Forecasting Dashboard")

# Introductory explanation
st.markdown("""
## Welcome to the Budgeting and Forecasting Dashboard

This app helps you create financial templates or forecast future expenses/revenue based on your data.

### How to Use This App:
1. **Choose an Action**: Select "Create a New Financial Template" to generate a template for data entry, or "Upload an Existing Financial Template" to forecast based on your data.
2. **Create a Template**: Pick your industry, set a date range, customize metrics, and download a template (CSV or Excel) to fill in your data. Enter only numbers in the template (negative values are allowed), without spaces, symbols (e.g., $, %), or commas.
3. **Forecast Data**: Upload your filled template, select your industry, specify the forecast horizon, and view/download the expanded dataset with forecasts. Note: A 'Total' column will be calculated from your metrics (e.g., sum of costs).
""")

# Button to reference instructions
with st.expander("View Instructions"):
    st.markdown("""
    ### How to Use This App:
    1. **Choose an Action**: Select "Create a New Financial Template" to generate a template for data entry, or "Upload an Existing Financial Template" to forecast based on your data.
    2. **Create a Template**: Pick your industry, set a date range, customize metrics, and download a template (CSV or Excel) to fill in your data. Enter only numbers in the template (negative values are allowed), without spaces, symbols (e.g., $, %), or commas.
    3. **Forecast Data**: Upload your filled template, select your industry, specify the forecast horizon, and view/download the expanded dataset with forecasts. Note: A 'Total' column will be calculated from your metrics (e.g., sum of costs).
    """)

# Button to explain how forecasting works
with st.expander("How Forecasting Works"):
    st.markdown("""
    ### How Forecasting Works:
    The app uses Prophet, a time-series forecasting model, to predict future values for each metric in your data (e.g., costs, revenue) based on historical trends. Grok analyzes your metrics and provides market trend adjustments (e.g., a percentage increase) to refine the forecast, ensuring predictions reflect both your data and industry trends.
    """)

# Initial prompt for user action
st.subheader("Get Started")
action = st.radio(
    "What would you like to do?",
    ["Create a New Financial Template", "Upload an Existing Financial Template"],
    help="Choose to either generate a new template for data entry or upload a completed template for forecasting."
)

# Handle "Create a New Financial Template" workflow
if action == "Create a New Financial Template":
    # Industry selection
    industry = st.selectbox("Select Your Industry Sector", list(INDUSTRY_FIELDS.keys()))

    # Date range selection
    st.subheader("Specify Your Date Range")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", value=datetime(2024, 1, 1), min_value=datetime(2000, 1, 1), max_value=datetime(2025, 12, 31))
    with col2:
        end_date = st.date_input("End Date", value=datetime(2025, 3, 31), min_value=start_date, max_value=datetime(2025, 12, 31))

    # Generate monthly dates between start and end
    dates = []
    current_date_loop = start_date
    while current_date_loop <= end_date:
        dates.append(current_date_loop.strftime("%Y-%m-%d"))
        current_date_loop += relativedelta(months=1)

    # Predefined fields based on industry (no Amount)
    predefined_fields = ["Date"] + INDUSTRY_FIELDS[industry]

    # Allow users to modify predefined fields (except Date)
    st.subheader("Customize Standard Financial Metrics")
    modified_fields = ["Date"]  # Date is mandatory
    for field in predefined_fields[1:]:  # Skip Date
        new_field_name = st.text_input(f"Rename or keep '{field}'", value=field, key=f"predefined_{field}")
        if new_field_name:  # Only add if not empty (i.e., not removed)
            modified_fields.append(new_field_name)

    # Allow users to add custom fields
    st.subheader("Include Additional Custom Metrics")
    custom_fields = []
    if 'custom_field_count' not in st.session_state:
        st.session_state.custom_field_count = 0

    if st.button("Add Custom Metric"):
        st.session_state.custom_field_count += 1

    for i in range(st.session_state.custom_field_count):
        custom_field = st.text_input(f"Custom Metric {i+1}", key=f"custom_{i}")
        if custom_field:
            custom_fields.append(custom_field)

    # Combine all fields
    all_fields = modified_fields + custom_fields

    # Choose download format
    download_format = st.selectbox("Choose Download Format", ["CSV", "Excel"])

    # Generate DataFrame for the file
    if dates and all_fields:
        df = pd.DataFrame(index=range(len(dates)), columns=all_fields)
        df['Date'] = dates
        # Leave other columns as NaN (empty) instead of 0 for cleaner input
        for col in all_fields[1:]:  # Skip Date
            df[col] = pd.NA

        st.subheader("Preview Your Financial Template")
        st.dataframe(df)

        # Generate file for download based on format
        file_name_base = f"budget_forecast_{industry.lower().replace(' ', '_')}_{current_date}"

        if download_format == "CSV":
            metadata = f"# Generated by Budgeting and Forecasting Dashboard on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} for {industry}"
            csv_buffer = io.StringIO()
            csv_buffer.write(metadata + "\n")
            df.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue()
            st.download_button(
                label="Download CSV Template",
                data=csv_data,
                file_name=f"{file_name_base}.csv",
                mime="text/csv"
            )

        else:  # Excel format
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Budget Data')
                workbook = writer.book
                worksheet = writer.sheets['Budget Data']
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
                worksheet.insert_rows(1)
                worksheet.cell(row=1, column=1).value = f"Generated by Budgeting and Forecasting Dashboard on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} for {industry}"
                worksheet.cell(row=1, column=1).font = Font(italic=True)
            excel_data = excel_buffer.getvalue()
            st.download_button(
                label="Download Excel Template",
                data=excel_data,
                file_name=f"{file_name_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("Please ensure a valid date range and at least one metric are selected.")

# Handle "Upload an Existing Financial Template" workflow
else:
    # Forecast horizon selection (moved inside this workflow)
    st.subheader("Forecasting Horizon")
    forecast_horizon = st.number_input(
        "How many months ahead would you like to forecast?",
        min_value=1,
        max_value=36,
        value=12,
        step=1,
        help="Specify the number of months for your forecast (e.g., 12 months for a year)."
    )

    st.subheader("Upload Your Completed Financial Template")
    st.info(
        "Please ensure your uploaded template follows the required format:\n"
        "- Columns must include 'Date' (format: YYYY-MM-DD, e.g., 2025-01-31) and your industry metrics.\n"
        "- Enter only numbers in the template (negative values are allowed), without spaces, symbols (e.g., $, %), or commas; the app will clean any such characters.\n"
        "- Ensure dates are consecutive months for accurate forecasting.\n"
        "- A 'Total' column will be calculated from your metrics (e.g., sum of costs)."
    )

    # Industry selection for upload
    uploaded_industry = st.selectbox("Select Your Industry Sector for Forecasting", list(INDUSTRY_FIELDS.keys()))

    uploaded_file = st.file_uploader("Choose your CSV or Excel file", type=["csv", "xlsx"])
    if uploaded_file:
        # Read the uploaded file, skipping comment lines starting with '#'
        try:
            if uploaded_file.name.endswith('.csv'):
                # Read the file as text to count comment lines
                uploaded_file.seek(0)  # Reset file pointer to start
                lines = uploaded_file.read().decode('utf-8').splitlines()
                skip_rows = 0
                for line in lines:
                    if line.startswith('#'):
                        skip_rows += 1
                    else:
                        break
                uploaded_file.seek(0)  # Reset file pointer again
                df = pd.read_csv(uploaded_file, skiprows=skip_rows)
            else:
                # Excel files don't typically have comment lines, but we'll check the first row
                df = pd.read_excel(uploaded_file)
                if df.columns[0].startswith('#'):
                    df = pd.read_excel(uploaded_file, skiprows=1)
        except Exception as e:
            st.error(f"Error reading the file: {str(e)}")
            st.stop()

        # Validate mandatory columns (removed Amount requirement)
        required_columns = ['Date']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"Missing required columns: {', '.join(missing_columns)}. Please ensure your file includes 'Date'.")
            st.stop()

        # Clean and standardize data
        try:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            for col in df.columns:
                if col != 'Date':
                    # Convert to string, strip spaces and non-numeric characters except negatives and decimals
                    df[col] = df[col].astype(str).str.replace(r'[^\d.-]', '', regex=True)
                    # Convert to numeric, preserving negatives
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            df.dropna(subset=['Date'], inplace=True)
            df.sort_values('Date', inplace=True)
        except Exception as e:
            st.error(f"Error processing data: {str(e)}. Ensure 'Date' is in YYYY-MM-DD format and metrics are numeric.")
            st.stop()

        # Check for sufficient data points
        if len(df) < 2:
            st.error("Insufficient data points for forecasting. Please provide at least 2 months of data.")
            st.stop()

        # Extract metrics (all columns except Date)
        metrics = [col for col in df.columns if col not in ['Date']]

        # Use Grok 3 Beta to analyze metrics
        st.subheader("Analyzing Your Financial Metrics")
        metrics_list = ", ".join(metrics)
        sample_data_dict = df[metrics].head(3).to_dict()
        regressor_prompt = (
            f"Given the following financial metrics for a {uploaded_industry} business: {metrics_list}, "
            f"with sample data: {sample_data_dict}, "
            f"identify which metrics are most likely to influence the overall financial forecast. "
            f"Provide a concise 1-2 sentence TLDR summary of your reasoning, ensuring each sentence is complete."
        )
        regressor_cache_key = f"regressor_analysis_{uploaded_industry}_{'_'.join(sorted(metrics))}"
        regressor_analysis = call_grok_api(regressor_prompt, regressor_cache_key)
        st.write("**Key Metrics Impacting Forecast**:", regressor_analysis)

        # Use Grok 3 Beta to get market trend adjustments for each metric individually
        st.subheader("Market Trend Adjustments")
        adjustments = {}
        for metric in metrics:
            trend_prompt = (
                f"For a {uploaded_industry} business, provide a concise 1-2 sentence TLDR summary of current market trends as of April 23, 2025, "
                f"that could specifically impact the metric '{metric}'. Suggest a percentage adjustment (e.g., increase or decrease) for this metric "
                f"over the next {forecast_horizon} months based on trends, ensuring each sentence is complete."
            )
            trend_cache_key = f"trend_adjust_{uploaded_industry}_{metric}_{forecast_horizon}"
            trend_analysis = call_grok_api(trend_prompt, trend_cache_key)
            st.write(f"**Trend for {metric}**:", trend_analysis)
            # Extract percentage adjustment (simplified parsing for prototype)
            try:
                if "increase by" in trend_analysis.lower():
                    percentage_str = trend_analysis.lower().split("increase by")[1].split("%")[0].strip()
                    adjustments[metric] = float(percentage_str)
                elif "decrease by" in trend_analysis.lower():
                    percentage_str = trend_analysis.lower().split("decrease by")[1].split("%")[0].strip()
                    adjustments[metric] = -float(percentage_str)
                else:
                    adjustments[metric] = 0.0  # Default to no adjustment if unclear
            except Exception as e:
                logger.warning(f"Could not parse percentage adjustment for {metric}: {str(e)}")
                adjustments[metric] = 0.0

        # Step 1: Forecast each metric individually using Prophet
        forecasted_metrics = {}
        for metric in metrics:
            # Prepare data for Prophet
            metric_df = df[['Date', metric]].rename(columns={'Date': 'ds', metric: 'y'})
            metric_df['ds'] = pd.to_datetime(metric_df['ds'])

            # Initialize Prophet model for the metric
            try:
                model = Prophet(yearly_seasonality=False, monthly_seasonality=False, weekly_seasonality=False, random_seed=42, changepoint_prior_scale=0.1)
            except TypeError:
                model = Prophet(yearly_seasonality=False, weekly_seasonality=False, changepoint_prior_scale=0.1)

            # Fit the model
            model.fit(metric_df)

            # Create future DataFrame
            future = model.make_future_dataframe(periods=forecast_horizon, freq='M')
            historical_dates = metric_df['ds'].dt.strftime('%Y-%m-%d').tolist()
            all_dates = future['ds'].dt.strftime('%Y-%m-%d').tolist()

            # Predict
            forecast = model.predict(future)

            # Apply percentage increase for future dates if the metric is in adjustments
            if metric in adjustments:
                percentage_increase = adjustments[metric]
                future_values = forecast['yhat'].values
                future_start_idx = len(historical_dates)
                for i in range(future_start_idx, len(all_dates)):
                    future_values[i] *= (1 + percentage_increase / 100)
                forecast['yhat'] = future_values

            # Store the forecasted values
            forecasted_metrics[metric] = forecast[['ds', 'yhat']].rename(columns={'ds': 'Date', 'yhat': metric})

        # Step 2: Combine forecasted metrics into a single DataFrame
        combined_forecast = None
        for metric, forecast_df in forecasted_metrics.items():
            if combined_forecast is None:
                combined_forecast = forecast_df
            else:
                combined_forecast = combined_forecast.merge(forecast_df[['Date', metric]], on='Date', how='left')

        # Step 3: Append forecasted rows to the original DataFrame
        combined_forecast['Date'] = pd.to_datetime(combined_forecast['Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        forecast_rows = combined_forecast[combined_forecast['Date'] > df['Date'].max()]

        # Add a "Data Type" column to distinguish actuals from forecasts
        df['Data Type'] = 'Actual'
        forecast_rows['Data Type'] = 'Forecast'
        expanded_df = pd.concat([df, forecast_rows], ignore_index=True)

        # Step 4: Calculate a "Total" column (sum of costs, excluding revenue metrics like Room Revenue)
        revenue_metrics = ['Room Revenue', 'Sales Revenue', 'Patient Revenue', 'Online Sales Revenue', 'Tuition Revenue', 'Project Revenue', 'Donations Revenue', 'Rental Income']
        cost_columns = [col for col in expanded_df.columns if col not in ['Date', 'Data Type'] and col not in revenue_metrics]
        expanded_df['Total'] = expanded_df[cost_columns].sum(axis=1)

        # Step 5: Round numeric columns to two decimal places (money format)
        numeric_columns = [col for col in expanded_df.columns if col not in ['Date', 'Data Type']]
        for col in numeric_columns:
            expanded_df[col] = expanded_df[col].round(2)

        # Step 6: Reorder columns to make Data Type the last column
        cols = [col for col in expanded_df.columns if col != 'Data Type'] + ['Data Type']
        expanded_df = expanded_df[cols]

        # Log the expanded DataFrame
        logger.info(f"Expanded DataFrame:\n{expanded_df.to_string()}")

        # Display the expanded DataFrame
        st.subheader("Expanded Financial Data with Forecast")
        st.write("The table below includes your original data with forecasted values appended for the next 12 months. A 'Total' column has been calculated from your metrics.")
        st.dataframe(expanded_df)

        # Plot forecast for Total using Plotly
        st.write("**Total Forecast Visualization**")
        try:
            # Split data into actuals and forecasts
            actuals_df = expanded_df[expanded_df['Data Type'] == 'Actual'][['Date', 'Total']]
            forecasts_df = expanded_df[expanded_df['Data Type'] == 'Forecast'][['Date', 'Total']]

            # Combine for the full timeline
            total_forecast = expanded_df[['Date', 'Total']].copy()
            total_forecast['Date'] = pd.to_datetime(total_forecast['Date'])

            # Find the transition point (last actual date)
            transition_date = actuals_df['Date'].max()

            # Find the maximum forecasted Total value
            max_forecast_total = forecasts_df['Total'].max()
            max_forecast_date = forecasts_df[forecasts_df['Total'] == max_forecast_total]['Date'].iloc[0]

            # Create a Plotly figure
            fig = go.Figure()

            # Plot actuals (solid line, blue)
            fig.add_trace(go.Scatter(
                x=actuals_df['Date'],
                y=actuals_df['Total'],
                mode='lines+markers',
                name='Actual Total',
                line=dict(color='#1f77b4', width=2),
                marker=dict(size=8)
            ))

            # Plot forecasts (dashed line, green)
            fig.add_trace(go.Scatter(
                x=forecasts_df['Date'],
                y=forecasts_df['Total'],
                mode='lines+markers',
                name='Forecasted Total',
                line=dict(color='#2ca02c', width=2, dash='dash'),
                marker=dict(size=8)
            ))

            # Add vertical line at the transition point
            fig.add_vline(
                x=transition_date.timestamp() * 1000,  # Convert to milliseconds for Plotly
                line=dict(color='gray', dash='dash'),
                annotation_text="Forecast Starts",
                annotation_position="top",
                annotation=dict(font_size=12, font_color="gray")
            )

            # Add annotation for the maximum forecast value
            fig.add_annotation(
                x=max_forecast_date,
                y=max_forecast_total,
                text=f"Peak Forecast: ${max_forecast_total:,.2f}",
                showarrow=True,
                arrowhead=2,
                ax=20,
                ay=-30,
                font=dict(size=12, color="red")
            )

            # Update layout for better styling
            fig.update_layout(
                title="Total Forecast Over Time",
                xaxis_title="Date",
                yaxis_title="Total ($)",
                hovermode="x unified",
                showlegend=True,
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(size=12),
                xaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray',
                    tickformat='%b %Y'
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray',
                    tickformat='$,.2f'
                )
            )

            # Display the plot
            st.plotly_chart(fig, use_container_width=True)

        except Exception as e:
            st.error(f"Error generating forecast plot: {str(e)}")

        # Grok 3 Beta methodology explanation with high effort
        methodology_prompt = (
            f"Given a financial forecast for a {uploaded_industry} business with the following metrics: {metrics_list}, "
            f"and forecast results: {forecast_rows.to_dict()}, "
            f"explain in a concise 1-2 sentence TLDR summary how each metric influenced the forecast of 'Total'. "
            f"Include a brief note on the forecasting methodology, focusing on trends and external factors, ensuring each sentence is complete."
        )
        methodology_cache_key = f"methodology_{uploaded_industry}_{'_'.join(sorted(metrics))}_{forecast_horizon}"
        methodology_explanation = call_grok_api(methodology_prompt, methodology_cache_key, high_effort=True)
        st.write("**Forecast Summary**")
        st.write(methodology_explanation)

        # Download the expanded DataFrame
        download_format = st.selectbox("Choose Forecast Download Format", ["CSV", "Excel"])
        file_name_base = f"expanded_forecast_{uploaded_industry.lower().replace(' ', '_')}_{current_date}"

        if download_format == "CSV":
            csv_buffer = io.StringIO()
            expanded_df.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue()
            st.download_button(
                label="Download Expanded Data as CSV",
                data=csv_data,
                file_name=f"{file_name_base}.csv",
                mime="text/csv"
            )
        else:
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Write the DataFrame to Excel
                expanded_df.to_excel(writer, index=False, sheet_name='Expanded Data')
                workbook = writer.book
                worksheet = writer.sheets['Expanded Data']

                # Define styles
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")  # Dark blue header
                actual_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # Light gray for actuals
                forecast_fill = PatternFill(start_color="B3DFFA", end_color="B3DFFA", fill_type="solid")  # Light blue for forecasts
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Style the header row
                for col_num, column_title in enumerate(expanded_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border

                # Find the row where actuals end and forecasts begin
                actual_end_idx = len(df)

                # Style the data rows
                for row_idx in range(2, len(expanded_df) + 2):
                    for col_idx in range(1, len(expanded_df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if col_idx != 1:  # Skip Date column
                            cell.number_format = '#,##0.00'  # Money format with 2 decimal places
                        if row_idx - 2 < actual_end_idx:
                            cell.fill = actual_fill
                        else:
                            cell.fill = forecast_fill
                        cell.border = border

                # Insert a blank row between actuals and forecasts
                worksheet.insert_rows(actual_end_idx + 2)

                # Adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

            excel_data = excel_buffer.getvalue()
            st.download_button(
                label="Download Expanded Data as Excel",
                data=excel_data,
                file_name=f"{file_name_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("Please upload a file to proceed with forecasting analysis.")